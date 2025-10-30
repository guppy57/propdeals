from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import questionary
import csv
import requests

load_dotenv()

EXCEL_FILE_PATH = "./Real Estate Investing.xlsx"
PROPERTIES_CSV_PATH = "./properties.csv"

RENTCAST_HEADERS = {
  "accept": "application/json",
  "X-Api-Key": os.getenv("RENTCAST_KEY")
}

console = Console()
wb = load_workbook(EXCEL_FILE_PATH)

def get_rentcast_data(property_details, rent_comps):
  pass

def get_geocode(address):
  response = requests.get(
      "https://maps.googleapis.com/maps/api/geocode/json",
      params={
          "key": os.getenv("GOOGLE_KEY"),
          "address": address,
      },
  )

  data = response.json()
  location = data["results"][0]["geometry"]["location"]
  return location["lng"], location["lat"]

def get_solar_potential_data(address):
    lng, lat = get_geocode(address)
    response = requests.get(
        "https://solar.googleapis.com/v1/buildingInsights:findClosest",
        params={
            "location.latitude": lat,
            "location.longitude": lng,
            "requiredQuality": "HIGH",
            "key": os.getenv("GOOGLE_KEY"),
        },
    )
    data = response.json()

    # Get first viable financial analysis (panelConfigIndex != -1)
    financial = next(
        (
            fa
            for fa in data.get("solarPotential", {}).get("financialAnalyses", [])
            if fa.get("panelConfigIndex", -1) != -1
        ),
        None,
    )

    solar_potential = data.get("solarPotential", {})
    building_stats = solar_potential.get("buildingStats", {})
    whole_roof = solar_potential.get("wholeRoofStats", {})

    result = {
        # Capacity metrics
        "max_array_panels_count": solar_potential.get("maxArrayPanelsCount"),
        "max_array_area_m2": solar_potential.get("maxArrayAreaMeters2"),
        "max_sunshine_hours_per_year": solar_potential.get("maxSunshineHoursPerYear"),
        "building_area_m2": building_stats.get("areaMeters2"),
        "whole_roof_area_m2": whole_roof.get("areaMeters2"),
        "panel_capacity_watts": solar_potential.get("panelCapacityWatts"),
        # Financial metrics (if viable config exists)
        "cost_electricity_without_solar": None,
        "payback_years": None,
        "lifetime_savings": None,
        "present_value_savings_year20": None,
        "upfront_cost": None,
        "solar_percentage": None,
        "initial_ac_kwh_per_year": None,
        "net_metering_allowed": None,
    }

    if financial and "financialDetails" in financial:
        details = financial["financialDetails"]
        cash = financial.get("cashPurchaseSavings", {})

        result.update(
            {
                "cost_electricity_without_solar": int(
                    details.get("costOfElectricityWithoutSolar", {}).get("units", 0)
                ),
                "payback_years": cash.get("paybackYears"),
                "lifetime_savings": int(
                    cash.get("savings", {}).get("savingsLifetime", {}).get("units", 0)
                ),
                "present_value_savings_year20": int(
                    cash.get("savings", {})
                    .get("presentValueOfSavingsYear20", {})
                    .get("units", 0)
                ),
                "upfront_cost": int(cash.get("upfrontCost", {}).get("units", 0)),
                "solar_percentage": details.get("solarPercentage"),
                "initial_ac_kwh_per_year": details.get("initialAcKwhPerYear"),
                "net_metering_allowed": details.get("netMeteringAllowed"),
            }
        )
    return result

def get_walkscore_data(property_details):
    lng, lat = get_geocode(property_details["full_address"])

    response = requests.get(
        "https://api.walkscore.com/score",
        params={
            "format": "json",
            "transit": 1,
            "bike": 1,
            "wsapikey": os.getenv("WALKSCORE_KEY"),
            "address": property_details["full_address"],
            "lat": lat,
            "lon": lng,
        },
    )
    data = response.json()

    try:
      transit = data['transit']['score'] 
    except KeyError:
      transit = "NA"
    
    try:
      bike = data['bike']['score']
    except KeyError:
      bike = "NA"

    return data["walkscore"], transit, bike 

def get_value_estimate(property_details):
  response = requests.get(
    "https://api.rentcast.io/v1/avm/value",
    headers=RENTCAST_HEADERS,
    params={
      "address": property_details["full_address"],
      "propertyType": "Multi-Family",
      "squareFootage": property_details["square_ft"],
      "bedrooms": property_details["beds"],
      "bathrooms": property_details["baths"],
      "compCount": 15
    }
  )

  data = response.json()
  return data

def collect_property_details():
    full_address = questionary.text("Full address").ask()
    zillow_link = questionary.text("Zillow link").ask()
    purchase_price = questionary.text("Purchase price").ask()
    property_type = questionary.select(
        "Property type",
        choices=["Duplex", "Triplex", "Fourplex"],
    ).ask()
    beds = questionary.text("Bedrooms").ask()
    baths = questionary.text("Bathrooms").ask()
    square_ft = questionary.text("Square footage").ask()
    built_in = questionary.text("Year built").ask()

    address1 = full_address.split(",")[0]
    # city = full_address.split(", ")[1].replace(" ", "")
    # state = full_address.split(", ")[2][1:].split(" ")[0]
    # zip = full_address.split(", ")[2][1:].split(" ")[1]

    unit_conversion = {
      "Duplex": 2,
      "Triplex": 3,
      "Fourplex": 4
    }

    units = unit_conversion[property_type]

    return {
      "full_address": full_address,
      "zillow_link": zillow_link,
      "purchase_price": int(purchase_price),
      "address1": address1,
      "beds": int(beds),
      "baths": int(baths),
      "square_ft": int(square_ft),
      "built_in": int(built_in),
      "units": units
    }

def get_test_data():
    return {
        "full_address": "224 Ash Ave, Ames, IA 50014",
        "zillow_link": "https://zillow.com/something-here",
        "purchase_price": int("269000"),
        "address1": "224 Ash Ave",
        "beds": int("3"),
        "baths": int("2"),
        "square_ft": int("2450"),
        "built_in": int("1909"),
        "units": 3,
    }

def display_property_details(property_details):
  text = ""
  
  for key in property_details:
    text += f"{key}: {property_details[key]}\n"

  console.print(Panel(text, title="Property Details", title_align="center", padding=1))

def add_property_sheet(property_details):
    template = wb["template"]
    prop = wb.copy_worksheet(template)
    prop.title = property_details["address1"]
    prop["B1"] = property_details["full_address"]
    prop["B1"].hyperlink = property_details["zillow_link"] 
    prop["B2"] = property_details["units"]
    prop["B3"] = property_details["purchase_price"]
    prop["B4"] = property_details["square_ft"]
    prop["B5"] = property_details["beds"]
    prop["B6"] = property_details["baths"]
    prop["B8"] = property_details["built_in"]

def add_property_to_csv(property_details):
  with open(PROPERTIES_CSV_PATH, 'a', newline='') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow([
      property_details["address1"],
      property_details["full_address"],
      property_details["zillow_link"],
      property_details["purchase_price"],
      property_details["beds"],
      property_details["baths"],
      property_details["square_ft"],
      property_details["built_in"],
      property_details["units"]
    ])

def edit_master_sheet(property_details):
    master = wb["Master"]
    last_row = master.max_row
    new_row = last_row + 1
    master[f"A{new_row}"] = property_details["address1"] 

def collect_rent_comps(unit_count):
  rent_comparables = []
  units_compared = 0

  while units_compared < unit_count:
    beds = questionary.text("Bedrooms (0 for Studio)").ask()
    baths = questionary.text("Bathrooms").ask()
    num_units = questionary.text("Number of units like this", default="1").ask()
    rent = questionary.text("Rent").ask()
    
    for i in range(int(num_units)):
      rent_comp = {
        "unit_num": units_compared + i,
        "beds": beds,
        "baths": baths,
        "rent": rent
      }
      rent_comparables.append(rent_comp)
    units_compared += int(num_units)
  
  return rent_comparables

def display_rent_comps(rent_comps):
  table = Table(title="Rent Comparables")
  table.add_column("Unit #")
  table.add_column("Configuration", no_wrap=True)
  table.add_column("Estimated Rent")

  for comp in rent_comps:
    configuration = f"{comp["beds"]}-beds {comp["baths"]}-baths"
    table.add_row(str(comp["unit_num"] + 1), configuration, str(comp["rent"]))
  
  console.print(table)

def add_rent_comps_to_sheet(property_details, rent_comps, unit_living_in):
  prop = wb[property_details["address1"]]
  prop["B43"] = unit_living_in

  unit_to_row = {
    0: 39,
    1: 40,
    2: 41,
    3: 42
  }

  for comp in rent_comps:
    row = unit_to_row[comp["unit_num"]]
    prop[f"B{row}"] = f"{comp["beds"]}-beds {comp["baths"]}-baths" 
    prop[f"C{row}"] = int(comp["rent"])

def add_rent_comps_to_csv(property_details, rent_comps):
  with open('./rent_estimates.csv', 'a', newline='') as csvfile:
    writer = csv.writer(csvfile)
    for comp in rent_comps:
      writer.writerow([property_details["address1"], int(comp["unit_num"]) + 1, comp["beds"], comp["beds"], comp["rent"]])

def backfill_neighborhood_data():
  pass

# --------------------------------------------------------

def run_program():
  console.print(Panel("Let's add a new property to analyze"), style="bold red")
  proceed = False

  while not proceed:
    property_details = collect_property_details()
    display_property_details(property_details=property_details)
    proceed = questionary.confirm("Does everything look correct?").ask()
    if not proceed:
      console.print("Add the property details again", style="bold blue")

  add_property_sheet(property_details)
  add_property_to_csv(property_details)
  edit_master_sheet(property_details)

  proceed2 = False

  while not proceed2:
    unit_count = property_details["units"]
    console.print(Panel(f"Let's now add our rent comparables for this property.\nWe will add details for the {unit_count} units.", title="Rent Comparables"), style="bold red")
    rent_comps = collect_rent_comps(unit_count)
    display_rent_comps(rent_comps)
    proceed2 = questionary.confirm("Does everything look correct?").ask()
    if not proceed2:
      console.print("Add the rent comparables again", style="bold blue")

  available_units = []

  for i in rent_comps:
    available_units.append(f"Unit {i["unit_num"] + 1}")

  unit_living_in = questionary.select("Unit living in", choices=available_units).ask()
  add_rent_comps_to_sheet(property_details, rent_comps, unit_living_in)
  add_rent_comps_to_csv(property_details, rent_comps)

  wb.save(EXCEL_FILE_PATH)
  wb.close()

  console.print(Panel("Workbook is saved and closed!"), style="bold green")

if __name__ == "__main__":
  run_program()