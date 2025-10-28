from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from openpyxl import load_workbook
import questionary
import csv

EXCEL_FILE_PATH = "/Users/armaangupta/Documents/Repositories/propdeals/Real Estate Investing.xlsx"

console = Console()
wb = load_workbook(EXCEL_FILE_PATH)

def collect_property_details():
    full_address = questionary.text("Full address").ask()
    zillow_link = questionary.text("Zillow link").ask()
    purchase_price = questionary.text("Purchase price").ask()
    property_type = questionary.select(
        "Property type",
        choices=["Duplex", "Triplex", "Fourplex"],
    ).ask()
    beds = questionary.text("Bedrooms").ask()
    baths = questionary.text("Bathrooms").ask()
    square_ft = questionary.text("Square footage").ask()
    lot_size = questionary.text("Lot size", default="0").ask()
    built_in = questionary.text("Year built").ask()

    address1 = full_address.split(",")[0]
    # city = full_address.split(", ")[1].replace(" ", "")
    # state = full_address.split(", ")[2][1:].split(" ")[0]
    # zip = full_address.split(", ")[2][1:].split(" ")[1]

    unit_conversion = {
      "Duplex": 2,
      "Triplex": 3,
      "Fourplex": 4
    }

    units = unit_conversion[property_type]

    return {
      "full_address": full_address,
      "zillow_link": zillow_link,
      "purchase_price": int(purchase_price),
      "address1": address1,
      "beds": int(beds),
      "baths": int(baths),
      "square_ft": int(square_ft),
      "lot_size": int(lot_size),
      "built_in": int(built_in),
      "units": units
    }

def get_test_data():
    return {
        "full_address": "123 Fuller Ave, Des Moines, IA 50010",
        "zillow_link": "https://zillow.com/something-here",
        "purchase_price": int("269000"),
        "address1": "123 Fuller Ave",
        "beds": int("3"),
        "baths": int("2"),
        "square_ft": int("2450"),
        "lot_size": int("0"),
        "built_in": int("1909"),
        "units": 3,
    }

def display_property_details(property_details):
  text = ""
  
  for key in property_details:
    text += f"{key}: {property_details[key]}\n"

  console.print(Panel(text, title="Property Details", title_align="center", padding=1))

def add_property_sheet(property_details):
    template = wb["template"]
    prop = wb.copy_worksheet(template)
    prop.title = property_details["address1"]
    prop["B1"] = property_details["full_address"]
    prop["B1"].hyperlink = property_details["zillow_link"] 
    prop["B2"] = property_details["units"]
    prop["B3"] = property_details["purchase_price"]
    prop["B4"] = property_details["square_ft"]
    prop["B5"] = property_details["beds"]
    prop["B6"] = property_details["baths"]
    prop["B7"] = property_details["lot_size"]
    prop["B8"] = property_details["built_in"]

def edit_master_sheet(property_details):
    master = wb["Master"]
    last_row = master.max_row
    new_row = last_row + 1
    master[f"A{new_row}"] = property_details["address1"] 

def collect_rent_comps(unit_count):
  rent_comparables = []
  units_compared = 0

  while units_compared < unit_count:
    beds = questionary.text("Bedrooms").ask()
    baths = questionary.text("Bathrooms").ask()
    num_units = questionary.text("Number of units like this", default="1").ask()
    rent = questionary.text("Rent").ask()
    
    for i in range(int(num_units)):
      rent_comp = {
        "unit_num": units_compared + i,
        "beds": beds,
        "baths": baths,
        "rent": rent
      }
      rent_comparables.append(rent_comp)
    units_compared += int(num_units)
  
  return rent_comparables

def display_rent_comps(rent_comps):
  table = Table(title="Rent Comparables")
  table.add_column("Unit #")
  table.add_column("Configuration", no_wrap=True)
  table.add_column("Estimated Rent")

  for comp in rent_comps:
    configuration = f"{comp["beds"]}-beds {comp["baths"]}-baths"
    table.add_row(str(comp["unit_num"] + 1), configuration, str(comp["rent"]))
  
  console.print(table)

def add_rent_comps(property_details, rent_comps, unit_living_in):
  prop = wb[property_details["address1"]]
  prop["B43"] = unit_living_in

  unit_to_row = {
    0: 39,
    1: 40,
    2: 41,
    3: 42
  }

  for comp in rent_comps:
    row = unit_to_row[comp["unit_num"]]
    prop[f"B{row}"] = f"{comp["beds"]}-beds {comp["baths"]}-baths" 
    prop[f"C{row}"] = int(comp["rent"])

console.print(Panel("Let's add a new property to analyze"), style="bold red")
proceed = False

while not proceed:
  property_details = collect_property_details()
  display_property_details(property_details=property_details)
  proceed = questionary.confirm("Does everything look correct?").ask()
  if not proceed:
    console.print("Add the property details again", style="bold blue")

property_details = get_test_data()

add_property_sheet(property_details)
edit_master_sheet(property_details)

proceed2 = False

while not proceed2:
  unit_count = property_details["units"]
  console.print(Panel(f"Let's now add our rent comparables for this property.\nWe will add details for the {unit_count} units.", title="Rent Comparables"), style="bold red")
  rent_comps = collect_rent_comps(unit_count)
  display_rent_comps(rent_comps)
  proceed2 = questionary.confirm("Does everything look correct?").ask()
  if not proceed2:
    console.print("Add the rent comparables again", style="bold blue")

available_units = []

for i in rent_comps:
  available_units.append(f"Unit {i["unit_num"] + 1}")

unit_living_in = questionary.select("Unit living in", choices=available_units).ask()
add_rent_comps(property_details, rent_comps, unit_living_in)

wb.save(EXCEL_FILE_PATH)
wb.close()

console.print(Panel("Workbook is saved and closed!"), style="bold green")