from openpyxl import load_workbook
import csv

wb = load_workbook("./Real Estate Investing.xlsx", data_only=True)

sheet_names = wb.sheetnames

# Iterate through all sheets
for sheet_name in wb.sheetnames:
    if sheet_name in ["Master", "Assumptions", "template"]:
      continue

    s = wb[sheet_name]
    print(f"Working with sheet: {sheet_name}")

    unit_to_row = {
      1: 39,
      2: 40,
      3: 41,
      4: 42
    }

    for key in unit_to_row:
        beds = 0
        baths = 0

        config = s[f"B{unit_to_row[key]}"].value
        estimate = s[f"C{unit_to_row[key]}"].value

        print(config)
        print(estimate)

        if config == "None":
          continue

        if config == "Studio":
          baths = 1
        else:
          beds_s = config.split(" ")[0]
          baths_s = config.split(" ")[1]
          beds = int(beds_s.split("-")[0])
          baths = int(baths_s.split("-")[0])
        
        with open('./rent_estimates.csv', 'a', newline='') as csvfile:
          writer = csv.writer(csvfile)
          writer.writerow([sheet_name, key, beds, baths, int(estimate)])

